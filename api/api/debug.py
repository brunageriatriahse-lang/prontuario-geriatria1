from http.server import BaseHTTPRequestHandler
import json, os, sys, traceback


class handler(BaseHTTPRequestHandler):

    def do_GET(self):
        resultado = {}

        # 1. Onde a função está rodando
        resultado["__file__"] = __file__
        resultado["cwd"] = os.getcwd()
        resultado["dirname"] = os.path.dirname(__file__)

        # 2. Arquivos disponíveis na pasta api/
        base = os.path.dirname(__file__)
        try:
            resultado["arquivos_api"] = os.listdir(base)
        except Exception as e:
            resultado["arquivos_api_erro"] = str(e)

        # 3. modelo.xlsm existe?
        modelo = os.path.join(base, "modelo.xlsm")
        resultado["modelo_existe"] = os.path.exists(modelo)
        if os.path.exists(modelo):
            resultado["modelo_tamanho"] = os.path.getsize(modelo)

        # 4. modelo_b64.txt existe?
        b64 = os.path.join(base, "modelo_b64.txt")
        resultado["b64_existe"] = os.path.exists(b64)

        # 5. openpyxl instalado?
        try:
            import openpyxl
            resultado["openpyxl"] = openpyxl.__version__
        except Exception as e:
            resultado["openpyxl_erro"] = str(e)

        # 6. Tenta carregar o modelo
        if os.path.exists(modelo):
            try:
                import io
                from openpyxl import load_workbook
                with open(modelo, "rb") as f:
                    dados = f.read()
                wb = load_workbook(io.BytesIO(dados), keep_vba=True)
                resultado["abas"] = wb.sheetnames
                resultado["carregamento"] = "OK"
            except Exception as e:
                resultado["carregamento_erro"] = str(e)
                resultado["carregamento_trace"] = traceback.format_exc()

        body = json.dumps(resultado, indent=2, ensure_ascii=False).encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, fmt, *args):
        pass
