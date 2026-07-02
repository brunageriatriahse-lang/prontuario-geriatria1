from http.server import BaseHTTPRequestHandler
import json, io, os, base64, traceback


def _carregar_modelo():
    base = os.path.dirname(__file__)
    p = os.path.join(base, "modelo.xlsm")
    if os.path.exists(p):
        with open(p, "rb") as f:
            return f.read()
    p2 = os.path.join(base, "modelo_b64.txt")
    if os.path.exists(p2):
        with open(p2, "r") as f:
            return base64.b64decode(f.read().strip())
    raise FileNotFoundError("modelo.xlsm nao encontrado em: " + base)


def _gerar(nome, prontuario, mae, idade, sexo, data):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(_carregar_modelo()), keep_vba=True)
    ws = wb["Cadastro"]
    ws["C7"].value = (nome or "").upper()
    ws["C8"].value = str(prontuario or "")
    ws["C9"].value = (mae or "").upper()
    ws["C12"].value = int(idade) if str(idade).strip().isdigit() else (idade or "")
    ws["C13"].value = sexo or ""
    ws["C14"].value = "GERIATRIA"
    ws["C16"].value = data or ""
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        try:
            n = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(n)) if n else {}
            arq = _gerar(
                body.get("nome", ""),
                body.get("prontuario", ""),
                body.get("maeNome", ""),
                body.get("idade", ""),
                body.get("sexo", ""),
                body.get("data", ""),
            )
            nome_arq = "Receituarios_" + (body.get("nome") or "paciente").replace(" ", "_") + ".xlsm"
            self.send_response(200)
            self._cors()
            self.send_header("Content-Type", "application/vnd.ms-excel.sheet.macroEnabled.12")
            self.send_header("Content-Disposition", 'attachment; filename="' + nome_arq + '"')
            self.send_header("Content-Length", str(len(arq)))
            self.end_headers()
            self.wfile.write(arq)
        except Exception as e:
            err = json.dumps({"error": str(e), "trace": traceback.format_exc()}).encode()
            self.send_response(500)
            self._cors()
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(err)))
            self.end_headers()
            self.wfile.write(err)

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def log_message(self, fmt, *args):
        pass
